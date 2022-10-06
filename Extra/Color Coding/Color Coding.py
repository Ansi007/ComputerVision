#!/usr/bin/env python
# coding: utf-8

# In[50]:


def colorCode():
    import pandas
    import numpy

    masterFile = pandas.read_excel('Master Input.xlsx')
    numberFile = pandas.read_excel('Numbers Input.xlsx')

    # print whole sheet data
    master = masterFile.to_numpy()
    number = numberFile.to_numpy()
    rowsMaster,colsMaster = master.shape
    rowsNumber,colsNumber = number.shape

    for i in range(0, rowsNumber):
        row = number[i]
        for a in range(0,colsMaster):
            for b in range(0,rowsMaster):
                row2 = []
                if(b+1) == 646:
                    continue
                if(str(master[b][a]) == 'nan'):
                    continue
                if(a+1) == 35:
                    continue
                if(str(master[b + 1][a]) == 'nan'):
                    if(str(master[b-24][a+1]) == 'nan'):
                        continue
                    if (b-24) < 0:
                        continue
                    row2 = [master[b][a], master[b-24][a+1]]
                    I1 = b
                    I2 = a
                    I3 = b-24
                    I4 = a+1
                else:
                    row2 = [master[b][a], master[b+1][a]]
                    I1 = b
                    I2 = a
                    I3 = b+1
                    I4 = a
                if(str(row[0]) == str(row2[0]) and str(row[1]) == str(row2[1])):
                    master[I1][I2] = str(master[I1][I2]) + 'ok'
                    master[I3][I4] = str(master[I3][I4]) + 'ok'

    outputArr = numpy.array(master)
    df = pandas.DataFrame(outputArr)
    filepath = 'OutputMasterTemporary.xlsx'
    df.to_excel(filepath, index=False)
    
    import openpyxl 
    from openpyxl.styles import PatternFill

    wb = openpyxl.load_workbook("OutputMasterTemporary.xlsx")
    ws = wb['Sheet1'] #Name of the working sheet

    fill_cell1 = PatternFill(patternType='solid', 
                               fgColor = '92D050')

    for row in ws:
        for cell in row:
            if 'ok' in str(cell.value):
                if 'nan' in str(cell.value):
                    cell.value = ''
                    continue
                cell.fill = fill_cell1
                value = str(cell.value)
                value = value.replace('ok','')
                cell.value = value

    wb.save("FinalOutput.xlsx")
    from os import remove 
    remove("OutputMasterTemporary.xlsx")
    showinfo(title = "Done!",message = "File Created Successfully")


# In[51]:


import tkinter as tk
from tkinter import ttk
from tkinter import filedialog as fd
from tkinter.messagebox import showinfo

# create the root window
root = tk.Tk()
root.title('Excel Scripting')
root.resizable(False, False)
root.geometry('200x50')

def callExcel():
    colorCode()

B = tk.Button(root, text ="Start", command = callExcel)

B.pack()

# run the application
root.mainloop()


# In[ ]:




